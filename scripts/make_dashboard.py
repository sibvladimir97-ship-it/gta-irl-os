#!/usr/bin/env python3
"""
GTA IRL OS — Dashboard Builder
Первый лист Excel — сводная панель управления.
"""
import json, glob, re, os, statistics
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

ROOT = '/Volumes/media/gta-irl-os'

# ── Цвета ─────────────────────────────────────────────────────────────────────
C = {
    'bg':       '0F172A',
    'surface':  '1E293B',
    'surface2': '334155',
    'blue':     '378ADD',
    'green':    '1D9E75',
    'amber':    'BA7517',
    'red':      'A32D2D',
    'purple':   '7F77DD',
    'teal':     '5DCAA5',
    'white':    'F0F0EE',
    'muted':    '64748B',
    'row_alt':  'F8FAFC',
    'row_hi':   'DCFCE7',
    'row_urg':  'FEF9C3',
}

def fill(color): return PatternFill('solid', start_color=color, end_color=color)
def font(color='F0F0EE', size=10, bold=False): return Font(name='Arial', size=size, bold=bold, color=color)
def align(h='left', v='center', wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border():
    t = Side(border_style='thin', color='1E293B')
    return Border(left=t, right=t, top=t, bottom=t)

MONTAGE_RE = re.compile(r"монтаж|видеомонтаж|монтажёр|монтажер|reels|рилс|шортс|shorts|premiere|capcut|davinci|after.effects|нарезка|субтитры", re.IGNORECASE)
BUDGET_RE  = re.compile(r"(\d[\d\s]{1,})\s*(?:тыс|тысяч)\b|(\d+)[кk]\b|(\d[\d\s]*)\s*(?:руб|₽)", re.IGNORECASE)
URGENT_RE  = re.compile(r"срочно|сегодня|asap|горит|немедленно", re.IGNORECASE)
FAST_RE    = re.compile(r"завтра|за \d+ (день|дня)|до пятницы|до понедельника|эта неделя", re.IGNORECASE)

NICHE_PAT = {
    'монтаж':   r"монтаж|reels|рилс|шортс|capcut|premiere|davinci",
    'telegram': r"telegram.?bot|телеграм.?бот|нужен.?бот|парсер",
    'ai':       r"\bai\b|нейросет|gpt|искусственный интеллект",
    'python':   r"python|автоматизация|скрипт",
    'smm':      r"smm|таргет|instagram|тикток|контент",
    'дизайн':   r"дизайн|логотип|баннер",
    'копир':    r"копирайт|тексты|статья",
    'сайт':     r"сайт|лендинг|web|приложен",
}

def extract_budget(text):
    for m in BUDGET_RE.finditer(text):
        raw = next(g for g in m.groups() if g)
        clean = re.sub(r'\s','',raw)
        try:
            v = int(clean)
            if 'тыс' in m.group(0) or 'тысяч' in m.group(0): v *= 1000
            if re.search(r'\dk|\dк', m.group(0), re.IGNORECASE): v *= 1000
            if 500 <= v <= 10_000_000: return v
        except: pass
    return None

def detect_niche(text):
    t = text.lower()
    for niche, pat in NICHE_PAT.items():
        if re.search(pat, t): return niche
    return 'другое'

def dedup(offers):
    seen = set()
    out = []
    for o in offers:
        key = f"{o.get('raw',{}).get('sender_id','')}:{o.get('raw',{}).get('msg_id','')}"
        if key not in seen:
            seen.add(key)
            out.append(o)
    return out

# ── Загрузка данных ───────────────────────────────────────────────────────────
files = glob.glob(f'{ROOT}/data/offers/*.json')
all_offers = []
for f in files:
    try: all_offers.append(json.load(open(f)))
    except: pass

raw_count  = len(all_offers)
offers     = dedup(all_offers)
dup_count  = raw_count - len(offers)

budgets    = [b for o in offers if (b := extract_budget(o.get('raw_text','') or '')) ]
niches     = [detect_niche(o.get('raw_text','') or '') for o in offers]
is_montage = [MONTAGE_RE.search(o.get('raw_text','') or '') for o in offers]
is_urgent  = [URGENT_RE.search(o.get('raw_text','') or '') for o in offers]
is_fast    = [FAST_RE.search(o.get('raw_text','') or '') for o in offers]
usernames  = set(o.get('raw',{}).get('sender_username','') for o in offers if o.get('raw',{}).get('sender_username'))

# Денежная статистика
total_budget  = sum(budgets)
avg_budget    = int(total_budget / len(budgets)) if budgets else 0
median_budget = int(statistics.median(budgets)) if budgets else 0
max_budget    = max(budgets) if budgets else 0
min_budget    = min(budgets) if budgets else 0

# Прибыль (оценки)
SELF_MARGIN     = 0.85  # выполнить самому — 85% маржа
DELEGATE_MARGIN = 0.30  # делегировать — 30% маржа
exp_self     = int(total_budget * SELF_MARGIN)
exp_delegate = int(total_budget * DELEGATE_MARGIN)

# Время до денег
can_today   = sum(1 for u in is_urgent if u)
can_3days   = sum(1 for f in is_fast if f)
can_week    = len([o for o in offers if not URGENT_RE.search(o.get('raw_text','') or '') and len((o.get('raw_text','') or '')) < 300])

# Ниши статистика
niche_stats = {}
for i, o in enumerate(offers):
    n = niches[i]
    b = extract_budget(o.get('raw_text','') or '')
    if n not in niche_stats:
        niche_stats[n] = {'count':0, 'budgets':[]}
    niche_stats[n]['count'] += 1
    if b: niche_stats[n]['budgets'].append(b)

# Монтаж отдельно
montage_offers = [o for i,o in enumerate(offers) if is_montage[i]]
montage_budgets = [b for o in montage_offers if (b := extract_budget(o.get('raw_text','') or ''))]
montage_avg    = int(sum(montage_budgets)/len(montage_budgets)) if montage_budgets else 0
montage_total  = sum(montage_budgets)

# Статусы (воронка)
status_map = {}
for o in all_offers:
    s = o.get('status','NEW')
    status_map[s] = status_map.get(s,0) + 1

# ── Excel ─────────────────────────────────────────────────────────────────────
# Загружаем существующий или создаём новый
xlsx_path = f'{ROOT}/gta_priority.xlsx'
if os.path.exists(xlsx_path):
    wb = load_workbook(xlsx_path)
    # Удаляем старый Dashboard если есть
    if '🏠 Dashboard' in wb.sheetnames:
        del wb['🏠 Dashboard']
    ws = wb.create_sheet('🏠 Dashboard', 0)
else:
    wb = Workbook()
    ws = wb.active
    ws.title = '🏠 Dashboard'

ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = '378ADD'

# Ширины колонок
col_widths = [2, 28, 18, 18, 18, 18, 2]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def cell(row, col, value='', bg=C['bg'], fg=C['white'], sz=10, bold=False,
         h='left', v='center', wrap=False, border=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(fg, sz, bold)
    c.alignment = align(h, v, wrap)
    if border:
        t = Side(border_style='thin', color='334155')
        c.border = Border(left=t, right=t, top=t, bottom=t)
    if num_fmt: c.number_format = num_fmt
    return c

def section_header(row, title, color=C['blue']):
    ws.merge_cells(f'B{row}:F{row}')
    c = ws.cell(row=row, column=2, value=f'  {title}')
    c.fill = fill(color)
    c.font = Font(name='Arial', size=11, bold=True, color=C['white'])
    c.alignment = align('left')
    ws.row_dimensions[row].height = 28

def kpi_row(row, label, value, value2=None, color=C['white'], fmt=None):
    ws.row_dimensions[row].height = 22
    c1 = ws.cell(row=row, column=2, value=f'  {label}')
    c1.fill = fill(C['surface'])
    c1.font = Font(name='Arial', size=10, color='94A3B8')
    c1.alignment = align('left')

    c2 = ws.cell(row=row, column=3, value=value)
    c2.fill = fill(C['surface'])
    c2.font = Font(name='Arial', size=12, bold=True, color=color)
    c2.alignment = align('right')
    if fmt: c2.number_format = fmt

    if value2 is not None:
        c3 = ws.cell(row=row, column=4, value=value2)
        c3.fill = fill(C['surface'])
        c3.font = Font(name='Arial', size=10, color='64748B')
        c3.alignment = align('left')

def spacer(row, height=8):
    ws.row_dimensions[row].height = height
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = fill(C['bg'])

def full_bg(row, height=20):
    ws.row_dimensions[row].height = height
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = fill(C['bg'])

# ── СТРОИМ DASHBOARD ─────────────────────────────────────────────────────────
r = 1

# Шапка
ws.row_dimensions[r].height = 14; full_bg(r); r += 1
ws.merge_cells(f'B{r}:F{r}')
c = ws.cell(row=r, column=2, value='  GTA IRL OS — Dashboard')
c.fill = fill(C['bg'])
c.font = Font(name='Arial', size=18, bold=True, color=C['white'])
c.alignment = align('left', 'center')
ws.row_dimensions[r].height = 40; r += 1

ws.merge_cells(f'B{r}:F{r}')
c = ws.cell(row=r, column=2, value=f'  Обновлено: {datetime.now().strftime("%d.%m.%Y %H:%M")}  ·  Данных: {len(offers)} уникальных офферов')
c.fill = fill(C['bg'])
c.font = Font(name='Arial', size=10, color='64748B')
c.alignment = align('left')
ws.row_dimensions[r].height = 20; r += 1
spacer(r); r += 1

# ── БЛОК 1: Общая статистика ──────────────────────────────────────────────────
section_header(r, '📊  Общая статистика'); r += 1
kpi_row(r, 'Всего найдено офферов',     raw_count,      color=C['blue']); r += 1
kpi_row(r, 'Уникальных (без дублей)',   len(offers),    color=C['teal']); r += 1
kpi_row(r, 'Дублей удалено',            dup_count,      color=C['muted']); r += 1
kpi_row(r, 'Уникальных заказчиков',     len(usernames), color=C['white']); r += 1
spacer(r); r += 1

# ── БЛОК 2: Деньги ────────────────────────────────────────────────────────────
section_header(r, '💰  Денежная статистика', C['green']); r += 1
kpi_row(r, 'Общий бюджет базы',          f'{total_budget:,} ₽'.replace(',', ' '),  color=C['green']); r += 1
kpi_row(r, 'Средний бюджет',             f'{avg_budget:,} ₽'.replace(',', ' '),    color=C['teal']); r += 1
kpi_row(r, 'Медианный бюджет',           f'{median_budget:,} ₽'.replace(',', ' '), color=C['teal']); r += 1
kpi_row(r, 'Максимальный бюджет',        f'{max_budget:,} ₽'.replace(',', ' '),    color=C['amber']); r += 1
kpi_row(r, 'Минимальный бюджет',         f'{min_budget:,} ₽'.replace(',', ' '),    color=C['muted']); r += 1
kpi_row(r, 'Прибыль (сам, 85% маржа)',   f'{exp_self:,} ₽'.replace(',', ' '),      color=C['green']); r += 1
kpi_row(r, 'Прибыль (делегировать, 30%)',f'{exp_delegate:,} ₽'.replace(',', ' '),  color=C['blue']); r += 1
spacer(r); r += 1

# ── БЛОК 3: Время до денег ────────────────────────────────────────────────────
section_header(r, '⏱  Время до получения денег', C['amber']); r += 1
kpi_row(r, 'Можно закрыть сегодня',   can_today,  value2='(срочные заказы)',   color=C['red']); r += 1
kpi_row(r, 'Можно закрыть за 3 дня',  can_3days,  value2='(с дедлайном)',      color=C['amber']); r += 1
kpi_row(r, 'Можно закрыть за неделю', can_week,   value2='(небольшие объёмы)', color=C['teal']); r += 1
spacer(r); r += 1

# ── БЛОК 4: Монтаж (1 очередь) ───────────────────────────────────────────────
section_header(r, '🎬  Монтаж — первая очередь', C['purple']); r += 1
kpi_row(r, 'Найдено заявок по монтажу',     len(montage_offers),                          color=C['purple']); r += 1
kpi_row(r, 'Средний чек',                   f'{montage_avg:,} ₽'.replace(',', ' '),       color=C['teal']); r += 1
kpi_row(r, 'Потенциальная прибыль (85%)',   f'{int(montage_total*0.85):,} ₽'.replace(',', ' '), color=C['green']); r += 1
kpi_row(r, 'Заявок с бюджетом',             len(montage_budgets),                         color=C['white']); r += 1
spacer(r); r += 1

# ── БЛОК 5: Ниши ─────────────────────────────────────────────────────────────
section_header(r, '🗂  Распределение по нишам', C['surface2']); r += 1

# Заголовки колонок ниш
niche_header_row = r
for col, title in [(2,'Ниша'), (3,'Кол-во'), (4,'Общий бюджет'), (5,'Средний чек'), (6,'Прибыль (85%)')]:
    c = ws.cell(row=r, column=col, value=title)
    c.fill = fill(C['surface2'])
    c.font = Font(name='Arial', size=9, bold=True, color='94A3B8')
    c.alignment = align('center' if col > 2 else 'left')
ws.row_dimensions[r].height = 20; r += 1

niche_order = ['монтаж','telegram','ai','python','smm','сайт','дизайн','копир','другое']
for niche in niche_order:
    if niche not in niche_stats: continue
    ns   = niche_stats[niche]
    tot  = sum(ns['budgets'])
    avg  = int(tot/len(ns['budgets'])) if ns['budgets'] else 0
    prof = int(tot * 0.85)
    icons = {'монтаж':'🎬','telegram':'🤖','ai':'🧠','python':'🐍','smm':'📱','сайт':'🌐','дизайн':'🎨','копир':'✍️','другое':'📦'}
    lbl = f"  {icons.get(niche,'•')} {niche}"

    row_bg = C['surface'] if niche_order.index(niche) % 2 == 0 else '172033'
    c1 = ws.cell(row=r, column=2, value=lbl)
    c1.fill = fill(row_bg); c1.font = Font(name='Arial', size=10, color=C['white']); c1.alignment = align('left')
    c2 = ws.cell(row=r, column=3, value=ns['count'])
    c2.fill = fill(row_bg); c2.font = Font(name='Arial', size=10, bold=True, color=C['blue']); c2.alignment = align('center')
    c3 = ws.cell(row=r, column=4, value=f'{tot:,} ₽'.replace(',', ' ') if tot else '—')
    c3.fill = fill(row_bg); c3.font = Font(name='Arial', size=10, color=C['teal']); c3.alignment = align('center')
    c4 = ws.cell(row=r, column=5, value=f'{avg:,} ₽'.replace(',', ' ') if avg else '—')
    c4.fill = fill(row_bg); c4.font = Font(name='Arial', size=10, color='94A3B8'); c4.alignment = align('center')
    c5 = ws.cell(row=r, column=6, value=f'{prof:,} ₽'.replace(',', ' ') if prof else '—')
    c5.fill = fill(row_bg); c5.font = Font(name='Arial', size=10, color=C['green']); c5.alignment = align('center')
    ws.row_dimensions[r].height = 20; r += 1

spacer(r); r += 1

# ── БЛОК 6: Воронка ───────────────────────────────────────────────────────────
section_header(r, '🔄  Воронка продаж', C['blue']); r += 1

funnel_stages = [
    ('Найдено',           raw_count),
    ('Уникальных',        len(offers)),
    ('Отклик отправлен',  status_map.get('RESPONDED', 0)),
    ('Ответили',          0),
    ('Переговоры',        0),
    ('Предоплата',        0),
    ('Закрыто',           0),
]

prev = raw_count
for stage, count in funnel_stages:
    conv = f'{round(count/prev*100)}%' if prev and count else '—'
    prev = count if count else prev
    c1 = ws.cell(row=r, column=2, value=f'  → {stage}')
    c1.fill = fill(C['surface']); c1.font = Font(name='Arial', size=10, color='94A3B8'); c1.alignment = align('left')
    c2 = ws.cell(row=r, column=3, value=count)
    c2.fill = fill(C['surface']); c2.font = Font(name='Arial', size=11, bold=True, color=C['blue']); c2.alignment = align('center')
    c3 = ws.cell(row=r, column=4, value=conv)
    c3.fill = fill(C['surface']); c3.font = Font(name='Arial', size=10, color='64748B'); c3.alignment = align('center')
    ws.row_dimensions[r].height = 20; r += 1

spacer(r); r += 1

# ── БЛОК 7: ТОП-10 прямо сейчас ──────────────────────────────────────────────
section_header(r, '🏆  ТОП-10 — открывать прямо сейчас', C['red']); r += 1

# Заголовки
for col, title in [(2,'#'), (3,'Ниша'), (4,'Контакт'), (5,'Бюджет'), (6,'Score')]:
    c = ws.cell(row=r, column=col, value=title)
    c.fill = fill(C['surface2']); c.font = Font(name='Arial', size=9, bold=True, color='94A3B8')
    c.alignment = align('center'); ws.row_dimensions[r].height = 18
r += 1

# Импортируем скоринг из priority_engine
import sys; sys.path.insert(0, f'{ROOT}/scripts')
try:
    from priority_engine import score_offer, deduplicate
    scored = [score_offer(o) for o in offers]
    scored.sort(key=lambda x: -x['priority'])
    top10 = scored[:10]
except:
    top10 = []

for i, s in enumerate(top10, 1):
    b = f'{s["budget"]:,}₽'.replace(',', ' ') if s.get('budget') else '?'
    row_bg = C['surface'] if i % 2 == 0 else '172033'

    c1 = ws.cell(row=r, column=2, value=i)
    c1.fill = fill(row_bg); c1.font = Font(name='Arial', size=10, bold=True, color=C['amber']); c1.alignment = align('center')

    c2 = ws.cell(row=r, column=3, value=s.get('niche',''))
    c2.fill = fill(row_bg); c2.font = Font(name='Arial', size=10, color=C['teal']); c2.alignment = align('center')

    uname = s.get('username','')
    c3 = ws.cell(row=r, column=4, value=f'@{uname}' if uname else '—')
    c3.fill = fill(row_bg); c3.font = Font(name='Arial', size=10, color=C['blue'], underline='single' if uname else 'none')
    if uname: c3.hyperlink = f'https://t.me/{uname}'
    c3.alignment = align('left')

    c4 = ws.cell(row=r, column=5, value=b)
    c4.fill = fill(row_bg); c4.font = Font(name='Arial', size=10, bold=True, color=C['green']); c4.alignment = align('center')

    c5 = ws.cell(row=r, column=6, value=s.get('priority',''))
    c5.fill = fill(row_bg); c5.font = Font(name='Arial', size=10, bold=True, color=C['white']); c5.alignment = align('center')

    ws.row_dimensions[r].height = 20; r += 1

spacer(r); r += 1

# ── БЛОК 8: AI статистика ─────────────────────────────────────────────────────
section_header(r, '🤖  AI & Эффективность', C['purple']); r += 1
kpi_row(r, 'Обработано без AI (rule-based)', f'{len(offers)} из {len(offers)}', value2='100%', color=C['green']); r += 1
kpi_row(r, 'AI-запросов на анализ',           0,  value2='(скоринг rule-based)', color=C['teal']); r += 1
kpi_row(r, 'Стоимость анализа 1 заявки',      '0₽', value2='(нет расхода токенов)', color=C['green']); r += 1
spacer(r); r += 1

# Заливаем фон колонок A и G
for row_i in range(1, r+5):
    for col_i in [1, 7]:
        ws.cell(row=row_i, column=col_i).fill = fill(C['bg'])
    for col_i in range(2, 7):
        if ws.cell(row=row_i, column=col_i).fill.fgColor.rgb in ('00000000', ''):
            ws.cell(row=row_i, column=col_i).fill = fill(C['bg'])

wb.save(xlsx_path)
print(f'Dashboard создан: {xlsx_path}')
print(f'Строк: {r}')

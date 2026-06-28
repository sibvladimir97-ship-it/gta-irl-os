import json, glob, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = '/Volumes/media/gta-irl-os'
files = glob.glob(f'{ROOT}/data/deals/active/*.json')
deals = []
for f in files:
    try:
        deals.append(json.load(open(f)))
    except:
        pass
deals.sort(key=lambda x: x.get('created_at',''), reverse=True)

wb = Workbook()
ws = wb.active
ws.title = 'Сделки'

hfill = PatternFill('solid', start_color='1A1A2E', end_color='1A1A2E')
hfont = Font(bold=True, color='FFFFFF', name='Arial', size=10)
altf  = PatternFill('solid', start_color='F8F9FA', end_color='F8F9FA')
lf    = Font(color='1A73E8', underline='single', name='Arial', size=10)
nf    = Font(name='Arial', size=10)
ca    = Alignment(horizontal='center', vertical='center')
la    = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin  = Side(border_style='thin', color='DEE2E6')
brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['#', 'Клиент', 'Telegram', 'Ниша / Заказ', 'Бюджет', 'Дедлайн', 'Статус', 'Последнее сообщение', 'Дата']
widths  = [4, 16, 22, 38, 12, 13, 24, 35, 12]

for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill
    cell.alignment = ca; cell.border = brd
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 26

STAGE_RU = {
    'NEW_LEAD':'Новый лид','RESPOND_DECIDED':'Решил откликнуться',
    'FIRST_MESSAGE_DRAFTED':'Черновик готов','FIRST_MESSAGE_SENT':'Отправлено',
    'WAITING_REPLY':'Ждём ответа','CLIENT_REPLIED':'Ответил',
    'QUALIFYING':'Квалификация','PROPOSAL_SENT':'КП отправлено',
    'PREPAYMENT_WAITING':'Ждём предоплату','PREPAYMENT_RECEIVED':'Предоплата',
    'IN_WORK':'В работе','CLOSED_WON':'Закрыто','CLOSED_LOST':'Потеряно',
    'CLIENT_GHOSTED':'Пропал','SCAM':'Скам',
}
SC = {
    'PREPAYMENT_RECEIVED':'C6EFCE','IN_WORK':'C6EFCE','CLOSED_WON':'C6EFCE',
    'WAITING_REPLY':'FFEB9C','QUALIFYING':'FFEB9C','PROPOSAL_SENT':'FCD5B0',
    'PREPAYMENT_WAITING':'FCD5B0','CLIENT_GHOSTED':'FFCCCC',
    'CLOSED_LOST':'FFCCCC','SCAM':'FFCCCC',
}

for i, d in enumerate(deals, 1):
    row = i + 1
    ct = d.get('contact', {})
    name = ct.get('name', '?')
    uname = ct.get('username', '')
    uid = str(ct.get('user_id', ''))
    stage = d.get('stage', '')
    budget = d.get('budget') or d.get('brief', {}).get('budget') or ''
    deadline = d.get('deadline') or d.get('brief', {}).get('deadline') or ''
    source = d.get('source', {}).get('chat', '')
    offer = (d.get('offer_text', '') or '')[:50]
    niche = f"{source}: {offer}" if offer else source
    created = (d.get('created_at', '') or '')[:10]
    msgs = d.get('messages', [])
    last = ''
    if msgs:
        m = msgs[-1]
        arrow = '>' if m['direction'] == 'outgoing' else '<'
        last = f"{arrow} {m['text'][:55]}"

    link = f'https://t.me/{uname}' if uname else (f'tg://user?id={uid}' if uid else '')
    disp = f'@{uname}' if uname else (f'ID:{uid}' if uid else '')

    fc = SC.get(stage)
    rf = PatternFill('solid', start_color=fc, end_color=fc) if fc else (altf if i%2==0 else None)

    def wr(col, val, font=None, align=None, hl=None):
        cl = ws.cell(row=row, column=col, value=val)
        cl.font = font or nf; cl.alignment = align or la; cl.border = brd
        if rf: cl.fill = rf
        if hl: cl.hyperlink = hl
        return cl

    wr(1, i, align=ca)
    wr(2, name)
    cl = wr(3, disp, font=lf if link else nf)
    if link: cl.hyperlink = link
    wr(4, niche)
    wr(5, budget, align=ca)
    wr(6, deadline, align=ca)
    wr(7, STAGE_RU.get(stage, stage), align=ca)
    wr(8, last)
    wr(9, created, align=ca)
    ws.row_dimensions[row].height = 20

ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(deals)+1}'
ws.freeze_panes = 'A2'

ws2 = wb.create_sheet('Сводка')
ws2['A1'] = 'GTA IRL OS — Сводка по сделкам'
ws2['A1'].font = Font(bold=True, size=13, name='Arial')
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 10
summary = {}
for d in deals:
    s = STAGE_RU.get(d.get('stage',''), d.get('stage',''))
    summary[s] = summary.get(s, 0) + 1
ws2['A3'] = 'Статус'; ws2['B3'] = 'Количество'
ws2['A3'].font = ws2['B3'].font = Font(bold=True, name='Arial')
for r, (s, n) in enumerate(sorted(summary.items()), 4):
    ws2.cell(row=r, column=1, value=s).font = Font(name='Arial', size=10)
    ws2.cell(row=r, column=2, value=n).font = Font(name='Arial', size=10)
tot = len(summary) + 4
ws2.cell(row=tot, column=1, value='ИТОГО').font = Font(bold=True, name='Arial')
ws2.cell(row=tot, column=2, value=f'=SUM(B4:B{tot-1})').font = Font(bold=True, name='Arial')

out = '/Volumes/media/gta-irl-os/gta_deals.xlsx'
wb.save(out)
print(f'OK: {len(deals)} сделок')

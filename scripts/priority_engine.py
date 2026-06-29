#!/usr/bin/env python3
"""
GTA IRL OS — Priority Engine v2
Полноценная CRM-таблица с дедупликацией, сортировкой, статусами, полным текстом.
Без LLM — только Python/regex/словари.
"""

import json
import glob
import re
import os
import hashlib
import requests
from datetime import datetime

ROOT      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BOT_TOKEN = os.getenv("TELEGRAM_TOKEN", "")
OWNER_ID  = 335699027

# ── Веса Priority Score (настраиваемые) ──────────────────────────────────────
# Формула: score = Σ(weight × factor_score)
# Факторы: budget, urgency, simplicity, clarity, close_prob
WEIGHTS = {
    "budget":     0.30,  # Высокий бюджет → быстрые деньги
    "urgency":    0.20,  # Срочность → клиент мотивирован платить сейчас
    "simplicity": 0.20,  # Простота → быстро сделать, высокая маржа
    "clarity":    0.15,  # Ясность ТЗ → меньше правок, меньше риска
    "close_prob": 0.15,  # Вероятность сделки → не тратить время впустую
}

# ── Ниши ─────────────────────────────────────────────────────────────────────

NICHE_PATTERNS = {
    "монтаж":       r"монтаж|видеомонтаж|монтажёр|монтажер|reels|рилс|шортс|shorts|premiere|capcut|davinci|after.effects|нарезка|субтитры",
    "youtube":      r"youtube|ютуб|ютьюб|ролик|видеоролик",
    "telegram_bot": r"telegram.?bot|телеграм.?бот|бот.?телеграм|нужен.?бот|парсер",
    "python":       r"python|питон|скрипт|автоматизация",
    "ai":           r"\bai\b|нейросет|gpt|искусственный интеллект|ai.агент|нейро",
    "smm":          r"smm|таргет|инстаграм|instagram|вконтакте|tiktok|тикток|контент.?мейкер",
    "копирайтинг":  r"копирайт|тексты|сео.тексты|seo.тексты|статья|описание.?товар",
    "дизайн":       r"дизайн|логотип|баннер|фотошоп|illustrator|figma",
    "разработка":   r"разработ|программист|сайт|лендинг|web|ios|android|приложен",
}

MONTAGE_RE = re.compile(
    r"монтаж|видеомонтаж|монтажёр|монтажер|reels|рилс|шортс|shorts|"
    r"premiere|capcut|davinci|after.effects|нарезка|субтитры|видеоредактор",
    re.IGNORECASE
)

# ── Бюджет ────────────────────────────────────────────────────────────────────

BUDGET_PATTERNS = [
    re.compile(r"(\d[\d\s]{2,})\s*(?:тыс|тысяч)\b", re.IGNORECASE),
    re.compile(r"(\d+)[кk]\b", re.IGNORECASE),
    re.compile(r"(\d[\d\s]*)\s*(?:руб|₽|rub)\b", re.IGNORECASE),
    re.compile(r"бюджет[:\s]+(\d[\d\s]*)", re.IGNORECASE),
    re.compile(r"оплата[:\s]+(\d[\d\s]*)", re.IGNORECASE),
    re.compile(r"от\s+(\d[\d\s]*)\s*(?:руб|₽)", re.IGNORECASE),
]

def extract_budget(text):
    for pat in BUDGET_PATTERNS:
        m = pat.search(text)
        if m:
            raw = m.group(1)
            clean = re.sub(r'\s', '', raw)
            try:
                val = int(clean)
                if "тыс" in pat.pattern or "тысяч" in pat.pattern:
                    val *= 1000
                if re.search(r"\dk\b|\dк\b", m.group(0), re.IGNORECASE):
                    val *= 1000
                if 500 <= val <= 10_000_000:
                    return val
            except:
                pass
    return None

def budget_score(budget):
    if not budget: return 3
    if budget >= 100000: return 10
    if budget >= 50000:  return 9
    if budget >= 20000:  return 7
    if budget >= 10000:  return 6
    if budget >= 5000:   return 4
    return 2

# ── Срочность ─────────────────────────────────────────────────────────────────

URGENT_HIGH = re.compile(r"срочно|сегодня|сейчас|asap|горит|немедленно|срочный", re.IGNORECASE)
URGENT_MED  = re.compile(r"завтра|на этой неделе|до пятницы|до понедельника|за \d+ (день|дня|дней)", re.IGNORECASE)
DEADLINE_RE = re.compile(r"до\s+\d{1,2}[./]\d{1,2}|\d{1,2}[./]\d{1,2}[./]\d{2,4}", re.IGNORECASE)

def urgency_level(text):
    if URGENT_HIGH.search(text):
        return "Высокая", 9
    if URGENT_MED.search(text) or DEADLINE_RE.search(text):
        return "Средняя", 6
    return "Низкая", 3

# ── Сложность ─────────────────────────────────────────────────────────────────

COMPLEX_STACK = re.compile(r"react|vue|angular|node|django|fastapi|docker|kubernetes|postgresql|redis", re.IGNORECASE)
NEED_TEAM     = re.compile(r"команд|несколько|2-3|3-5 чел|группа", re.IGNORECASE)

def complexity_score(text, niche):
    score = 5
    if niche in ("монтаж", "копирайтинг"):     score -= 2
    if niche in ("telegram_bot", "python"):     score += 0
    if niche in ("разработка",):                score += 2
    if COMPLEX_STACK.search(text):              score += 2
    if NEED_TEAM.search(text):                  score += 2
    req_count = len(re.findall(r"[\n•\-]\s*\w", text))
    if req_count > 10: score += 2
    elif req_count > 5: score += 1
    if len(text) > 1000: score += 1
    return max(1, min(10, score))

def complexity_label(score):
    if score <= 3: return "Легко"
    if score <= 6: return "Средне"
    return "Сложно"

# ── Ясность ТЗ ────────────────────────────────────────────────────────────────

def clarity_score(text):
    score = 3
    if len(text) > 500:  score += 2
    if len(text) > 200:  score += 1
    if re.search(r"\d", text): score += 1
    if re.search(r"нужно|требуется|необходимо|должен|хочу чтобы", text, re.IGNORECASE): score += 1
    if re.search(r"пример|образец|референс|как.например", text, re.IGNORECASE): score += 1
    if re.search(r"тз|техническое задание|подробн", text, re.IGNORECASE): score += 1
    return max(1, min(10, score))

# ── Вероятность закрытия ──────────────────────────────────────────────────────

POSITIVE_RE = re.compile(r"готов заплатить|оплачу|бюджет|срочно|нужен сегодня|ищу|нужна помощь", re.IGNORECASE)
NEGATIVE_RE = re.compile(r"за опыт|бесплатно|за отзыв|тест|студент|волонтёр", re.IGNORECASE)

def close_prob_score(text):
    score = 5
    score += min(3, len(POSITIVE_RE.findall(text)))
    score -= min(4, len(NEGATIVE_RE.findall(text)) * 2)
    return max(1, min(10, score))

# ── Ниша и услуга ────────────────────────────────────────────────────────────

SERVICE_NAMES = {
    "монтаж":       "Видеомонтаж",
    "youtube":      "YouTube-контент",
    "telegram_bot": "Telegram-бот",
    "python":       "Python/автоматизация",
    "ai":           "AI-агент/нейросети",
    "smm":          "SMM/контент",
    "копирайтинг":  "Копирайтинг",
    "дизайн":       "Дизайн",
    "разработка":   "Веб/приложение",
    "прочее":       "Другое",
}

def detect_niche(text):
    t = text.lower()
    for niche, pattern in NICHE_PATTERNS.items():
        if re.search(pattern, t):
            return niche
    return "прочее"

# ── Дедупликация ─────────────────────────────────────────────────────────────

def offer_fingerprint(offer):
    """Отпечаток для дедупликации."""
    text = (offer.get("raw_text","") or "")[:150]
    uid  = str(offer.get("raw",{}).get("sender_id",""))
    mid  = str(offer.get("raw",{}).get("msg_id",""))
    return hashlib.md5(f"{uid}:{mid}:{text}".encode()).hexdigest()

def deduplicate(offers):
    seen = set()
    result = []
    for o in offers:
        fp = offer_fingerprint(o)
        if fp not in seen:
            seen.add(fp)
            result.append(o)
    return result

# ── Форматирование даты ───────────────────────────────────────────────────────

def fmt_date(raw):
    if not raw:
        return ""
    for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(raw[:26], fmt)
            if dt.hour or dt.minute:
                return dt.strftime("%d.%m.%Y %H:%M")
            return dt.strftime("%d.%m.%Y")
        except:
            pass
    return raw[:10]

# ── Статусы ───────────────────────────────────────────────────────────────────

STATUS_WORKFLOW = [
    "Новое", "Изучено", "Отклик отправлен", "Ответ получен",
    "Жду ТЗ", "Получено ТЗ", "Переговоры", "Предоплата",
    "В работе", "Делегировано", "Готово", "Закрыто", "Отказ",
]

def map_status(raw_status):
    mapping = {
        "NEW":              "Новое",
        "RESPONDED":        "Отклик отправлен",
        "HIDDEN":           "Закрыто",
        "SCAM":             "Отказ",
        "DELEGATED":        "Делегировано",
    }
    return mapping.get(raw_status, "Новое")

# ── Основной скоринг ─────────────────────────────────────────────────────────

def score_offer(offer):
    text  = offer.get("raw_text","") or ""
    d     = offer.get("display",{})
    raw   = offer.get("raw",{})

    niche      = detect_niche(text)
    service    = SERVICE_NAMES.get(niche, "Другое")
    budget     = extract_budget(text)
    urgency_lbl, urgency_s = urgency_level(text)
    complex_s  = complexity_score(text, niche)
    clarity_s  = clarity_score(text)
    close_s    = close_prob_score(text)
    budget_s   = budget_score(budget)

    priority = round(
        budget_s  * WEIGHTS["budget"] +
        urgency_s * WEIGHTS["urgency"] +
        (10 - complex_s) * WEIGHTS["simplicity"] +  # простота = инверсия сложности
        clarity_s * WEIGHTS["clarity"] +
        close_s   * WEIGHTS["close_prob"],
        2
    )

    username = raw.get("sender_username") or ""
    contact_url = f"https://t.me/{username}" if username else ""
    msg_url = raw.get("msg_url","") or ""

    return {
        "priority":       priority,
        "id":             offer["offer_id"],
        "date":           fmt_date(offer.get("created_at","")),
        "source_link":    msg_url,
        "source_name":    d.get("chat_name",""),
        "username":       username,
        "contact_url":    contact_url,
        "niche":          niche,
        "service":        service,
        "budget":         budget or "",
        "urgency":        urgency_lbl,
        "urgency_score":  urgency_s,
        "complexity":     complexity_label(complex_s),
        "complexity_num": complex_s,
        "clarity":        clarity_s,
        "close_prob":     close_s,
        "self_do":        "Да" if niche in ("монтаж","telegram_bot","python","ai","youtube") else "Нет",
        "delegate":       "Да" if niche in ("монтаж","дизайн","копирайтинг","smm") else "Нет",
        "status":         map_status(offer.get("status","NEW")),
        "is_montage":     bool(MONTAGE_RE.search(text)),
        "preview":        (text[:150] + "...") if len(text) > 150 else text,
        "full_text":      text,
        # Priority Score breakdown для прозрачности
        "score_budget":   round(budget_s  * WEIGHTS["budget"], 2),
        "score_urgency":  round(urgency_s * WEIGHTS["urgency"], 2),
        "score_simplicity": round((10-complex_s) * WEIGHTS["simplicity"], 2),
        "score_clarity":  round(clarity_s * WEIGHTS["clarity"], 2),
        "score_close":    round(close_s   * WEIGHTS["close_prob"], 2),
    }

# ── Excel ─────────────────────────────────────────────────────────────────────

def make_excel(scored):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    hfill  = PatternFill('solid', start_color='0F172A', end_color='0F172A')
    hfont  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    nf     = Font(name='Arial', size=9)
    lf     = Font(color='1A73E8', underline='single', name='Arial', size=9)
    bf     = Font(name='Arial', size=9, bold=True)
    ca     = Alignment(horizontal='center', vertical='center')
    la     = Alignment(horizontal='left', vertical='center', wrap_text=False)
    lw     = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin   = Side(border_style='thin', color='E2E8F0')
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Цвет строки
    FILL_MONTAGE  = PatternFill('solid', start_color='DCFCE7', end_color='DCFCE7')
    FILL_HIGH     = PatternFill('solid', start_color='FEF9C3', end_color='FEF9C3')
    FILL_URGENT   = PatternFill('solid', start_color='FEE2E2', end_color='FEE2E2')
    FILL_ALT      = PatternFill('solid', start_color='F8FAFC', end_color='F8FAFC')

    def row_fill(s, idx):
        if s["is_montage"]:        return FILL_MONTAGE
        if s["urgency"] == "Высокая": return FILL_URGENT
        if s["priority"] >= 7:     return FILL_HIGH
        if idx % 2 == 0:           return FILL_ALT
        return None

    def write_header(ws, cols):
        for c, (h, w) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hfont; cell.fill = hfill
            cell.alignment = ca; cell.border = brd
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[1].height = 26

    def write_row(ws, row, s, idx, cols_def):
        rf = row_fill(s, idx)
        for c, (key, fmt) in enumerate(cols_def, 1):
            val = s.get(key, "")
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = brd
            if rf: cell.fill = rf
            if fmt == "link" and s.get(key + "_url"):
                cell.font = lf
                cell.hyperlink = s[key + "_url"]
                cell.alignment = la
            elif fmt == "center":
                cell.font = nf; cell.alignment = ca
            elif fmt == "bold":
                cell.font = bf; cell.alignment = ca
            elif fmt == "wrap":
                cell.font = nf; cell.alignment = lw
            else:
                cell.font = nf; cell.alignment = la
        ws.row_dimensions[row].height = 18

    # ── Главный лист ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "📋 Все офферы"

    main_cols = [
        ("Priority ↕", 9), ("ID", 9), ("Дата ↕", 12), ("Источник", 22),
        ("Контакт ↕", 18), ("Ниша ↕", 13), ("Услуга ↕", 18),
        ("Бюджет ↕", 10), ("Срочность ↕", 11), ("Сложность ↕", 11),
        ("Ясность ТЗ ↕", 10), ("Вероятность ↕", 10),
        ("Сам", 6), ("Делегировать", 10),
        ("Статус ↕", 16), ("Превью", 55),
    ]
    write_header(ws, main_cols)

    cols_def = [
        ("priority","bold"), ("id","center"), ("date","center"),
        ("source_name","link"), ("username","link"),
        ("niche","center"), ("service",""),
        ("budget","center"), ("urgency","center"), ("complexity","center"),
        ("clarity","center"), ("close_prob","center"),
        ("self_do","center"), ("delegate","center"),
        ("status","center"), ("preview",""),
    ]
    # Исправляем ссылки
    for s in scored:
        s["source_name_url"] = s["source_link"]
        s["username_url"]    = s["contact_url"]

    for i, s in enumerate(scored, 2):
        write_row(ws, i, s, i, cols_def)

    ws.auto_filter.ref = f'A1:{get_column_letter(len(main_cols))}{len(scored)+1}'
    ws.freeze_panes = 'A2'

    # ── Лист: Монтаж (1 очередь) ──────────────────────────────────────────────
    ws2 = wb.create_sheet("🎬 Монтаж (1 очередь)")
    montage = sorted([s for s in scored if s["is_montage"]], key=lambda x: -x["priority"])
    _write_crm_sheet(ws2, montage, write_header, write_row, main_cols, cols_def)

    # ── Лист: ТОП Priority ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("🏆 ТОП Priority")
    top = sorted(scored, key=lambda x: -x["priority"])[:50]
    _write_crm_sheet(ws3, top, write_header, write_row, main_cols, cols_def)

    # ── Лист: ТОП Бюджет ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("💰 ТОП Бюджет")
    by_budget = sorted([s for s in scored if s["budget"]], key=lambda x: -x["budget"])[:40]
    _write_crm_sheet(ws4, by_budget, write_header, write_row, main_cols, cols_def)

    # ── Лист: Срочные ─────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("🔥 Срочные")
    urgent = [s for s in scored if s["urgency"] == "Высокая"]
    _write_crm_sheet(ws5, urgent, write_header, write_row, main_cols, cols_def)

    # ── Лист: Полные тексты (для работы с клиентами) ─────────────────────────
    ws6 = wb.create_sheet("📄 Полные тексты")
    full_cols = [("Priority",8),("Ниша",12),("Контакт",18),("Бюджет",10),("Полный текст",80)]
    write_header(ws6, full_cols)
    top_full = sorted(scored, key=lambda x: -x["priority"])[:100]
    for i, s in enumerate(top_full, 2):
        ws6.cell(row=i, column=1, value=s["priority"]).font = Font(bold=True, name='Arial', size=9)
        ws6.cell(row=i, column=2, value=s["niche"]).font = Font(name='Arial', size=9)
        cl = ws6.cell(row=i, column=3, value=s["username"] or "—")
        cl.font = Font(color='1A73E8', underline='single', name='Arial', size=9) if s["contact_url"] else Font(name='Arial', size=9)
        if s["contact_url"]: cl.hyperlink = s["contact_url"]
        ws6.cell(row=i, column=4, value=s["budget"] or "?").font = Font(name='Arial', size=9)
        cl2 = ws6.cell(row=i, column=5, value=s["full_text"])
        cl2.font = Font(name='Arial', size=9)
        cl2.alignment = Alignment(wrap_text=True, vertical='top')
        ws6.row_dimensions[i].height = 60

    ws6.column_dimensions['E'].width = 80
    ws6.freeze_panes = 'A2'

    # ── Лист: Формула Priority Score ─────────────────────────────────────────
    ws7 = wb.create_sheet("⚙️ Формула")
    ws7['A1'] = 'Priority Score — Формула расчёта'
    ws7['A1'].font = Font(bold=True, size=13, name='Arial')
    ws7['A3'] = 'Priority Score = Σ (Вес × Оценка фактора)'
    ws7['A3'].font = Font(bold=True, name='Arial')
    rows = [
        ("Фактор", "Вес", "Макс. оценка", "Описание"),
        ("Бюджет",      f"{WEIGHTS['budget']:.0%}",     "10", "Чем выше бюджет — тем выше балл"),
        ("Срочность",   f"{WEIGHTS['urgency']:.0%}",    "9",  "Срочно/сегодня/ASAP → 9, дедлайн → 6, нет → 3"),
        ("Простота",    f"{WEIGHTS['simplicity']:.0%}", "10", "10 − сложность. Монтаж=легко, разработка=сложно"),
        ("Ясность ТЗ",  f"{WEIGHTS['clarity']:.0%}",   "10", "Длина текста + конкретика + примеры"),
        ("Вер. сделки", f"{WEIGHTS['close_prob']:.0%}", "10", "Наличие бюджета, срочности; минус 'за отзыв'"),
        ("ИТОГО",       "100%",                          "10", "Диапазон от 1 до 10"),
    ]
    for r, row_data in enumerate(rows, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws7.cell(row=r, column=c, value=val)
            if r == 5:
                cell.font = Font(bold=True, name='Arial', size=10)
                cell.fill = PatternFill('solid', start_color='0F172A', end_color='0F172A')
                cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            else:
                cell.font = Font(name='Arial', size=10)
    ws7.column_dimensions['A'].width = 18
    ws7.column_dimensions['B'].width = 8
    ws7.column_dimensions['C'].width = 14
    ws7.column_dimensions['D'].width = 50

    out = os.path.join(ROOT, "gta_priority.xlsx")
    wb.save(out)
    return out, len(montage)

def _write_crm_sheet(ws, data, write_header, write_row, main_cols, cols_def):
    write_header(ws, main_cols)
    for i, s in enumerate(data, 2):
        write_row(ws, i, s, i, cols_def)
    from openpyxl.utils import get_column_letter
    ws.auto_filter.ref = f'A1:{get_column_letter(len(main_cols))}{len(data)+1}'
    ws.freeze_panes = 'A2'

# ── Telegram отчёт ────────────────────────────────────────────────────────────

def send_report(scored, montage_count):
    if not BOT_TOKEN:
        return

    total       = len(scored)
    with_budget = [s for s in scored if s["budget"]]
    avg_budget  = int(sum(s["budget"] for s in with_budget) / len(with_budget)) if with_budget else 0
    urgent_cnt  = sum(1 for s in scored if s["urgency"] == "Высокая")

    top20 = sorted(scored, key=lambda x: -x["priority"])[:20]

    lines = [
        "📊 *Priority Engine v2 — Готово*", "",
        f"Обработано: *{total}* офферов",
        f"По монтажу (1 очередь): *{montage_count}*",
        f"Срочных: *{urgent_cnt}*",
        f"Средний бюджет: *{avg_budget:,} ₽*".replace(",", " "), "",
        "*ТОП-20 — открывай прямо сейчас:*",
    ]

    for i, s in enumerate(top20, 1):
        b = f"{s['budget']:,}₽".replace(",", " ") if s["budget"] else "?"
        c = f"@{s['username']}" if s["username"] else "нет контакта"
        u = s["urgency"][0] if s["urgency"] else "-"
        lines.append(f"{i}. [{s['niche']}] {b} · {c} · score:{s['priority']} · срочн:{u}")

    lines += ["", "📁 Файл: gta\\_priority.xlsx", "Листы: Все / Монтаж / ТОП / Срочные / Тексты / Формула"]

    requests.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
        json={"chat_id": OWNER_ID, "text": "\n".join(lines),
              "parse_mode": "Markdown", "disable_web_page_preview": True},
        timeout=15
    )

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("GTA IRL OS — Priority Engine v2")

    files = glob.glob(os.path.join(ROOT, "data", "offers", "*.json"))
    offers = []
    for f in files:
        try: offers.append(json.load(open(f)))
        except: pass

    print(f"Загружено: {len(offers)}")

    before = len(offers)
    offers = deduplicate(offers)
    print(f"После дедупликации: {len(offers)} (убрано {before - len(offers)} дублей)")

    scored = [score_offer(o) for o in offers]
    scored.sort(key=lambda x: -x["priority"])

    montage   = [s for s in scored if s["is_montage"]]
    with_b    = [s for s in scored if s["budget"]]
    avg_b     = int(sum(s["budget"] for s in with_b) / len(with_b)) if with_b else 0
    urgent    = [s for s in scored if s["urgency"] == "Высокая"]

    print(f"\nРезультат:")
    print(f"  Всего уникальных: {len(scored)}")
    print(f"  Монтаж: {len(montage)}")
    print(f"  Срочных: {len(urgent)}")
    print(f"  Средний бюджет: {avg_b:,} ₽")

    print("\nТОП-10 прямо сейчас:")
    for i, s in enumerate(scored[:10], 1):
        b = f"{s['budget']:,}₽" if s["budget"] else "?"
        print(f"  {i}. score:{s['priority']} [{s['niche']}] {b} @{s['username'] or '?'} | {s['urgency']} | {s['status']}")

    print("\nСоздаю Excel (6 листов)...")
    out, mc = make_excel(scored)
    print(f"Сохранено: {out}")

    print("Отправляю в Telegram...")
    send_report(scored, mc)
    print("Готово!")

if __name__ == "__main__":
    main()
